import io

import pytest
from openpyxl import load_workbook

from src.export.xlsx_exporter import markdown_to_xlsx


class TestMarkdownToXlsx:
    def test_basic_conversion(self):
        md = """\
| 番号 | サービス | 金額(円) | 備考 |
| --- | --- | --- | --- |
| 03-1234-5678 | 基本料 | 1800 | 2025年7月分 |
| 03-1234-5678 | ナンバーディスプレイ | 400 |  |
"""
        xlsx_bytes = markdown_to_xlsx(md)
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # ヘッダー行 + 2データ行 = 3行
        assert len(rows) == 3
        assert rows[0] == ("番号", "サービス", "金額(円)", "備考")
        assert rows[1][0] == "03-1234-5678"
        assert rows[1][1] == "基本料"
        assert rows[1][2] == "1800"

    def test_empty_raises(self):
        with pytest.raises(ValueError):
            markdown_to_xlsx("")

    def test_separator_only_raises(self):
        md = "| --- | --- | --- | --- |"
        with pytest.raises(ValueError):
            markdown_to_xlsx(md)
